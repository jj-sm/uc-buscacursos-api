import io
import re
from typing import List, Optional
from datetime import datetime

from openpyxl import load_workbook

try:
    import xlrd
except ImportError:
    xlrd = None

from .notam_store import insert_notam


EXPECTED_HEADERS = [
    "location",
    "notam #/lta #",
    "class",
    "issue date (utc)",
    "effective date (utc)",
    "expiration date (utc)",
    "notam condition/lta subject/construction graphic title",
]


def _norm(s: Optional[str]) -> str:
    if not s:
        return ""
    s = str(s)
    s = s.replace("\uFEFF", "")
    s = re.sub(r"[\x00-\x1F]+", " ", s)
    return re.sub(r"\s+", " ", s).strip().lower()


def _find_header_map(row_values: List[str]):
    """Return column index mapping {expected_index -> col_index} if match found."""
    row_norm = [_norm(v) for v in row_values]
    col_map = {}

    for exp_idx, exp_header in enumerate(EXPECTED_HEADERS):
        for col_idx, val in enumerate(row_norm):
            if exp_header == val:
                col_map[exp_idx] = col_idx
                break

    return col_map if len(col_map) >= 6 else None


def read_xlsx_rows(file_bytes: bytes) -> List[List[str]]:
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active

    header_row_idx = None
    col_map = None

    # 1️⃣ Find header row
    for row in ws.iter_rows(min_row=1, max_row=30):
        values = [c.value if c.value is not None else "" for c in row]
        col_map = _find_header_map(values)
        if col_map:
            header_row_idx = row[0].row
            break

    if header_row_idx is None:
        raise ValueError("Unable to locate NOTAM header row in XLSX file")

    # 2️⃣ Parse rows
    rows = []
    for row in ws.iter_rows(min_row=header_row_idx + 1):
        vals = []
        empty = True
        for exp_idx in range(7):
            col_idx = col_map.get(exp_idx)
            v = ""
            if col_idx is not None and col_idx < len(row):
                cell = row[col_idx].value
                if cell is not None:
                    v = str(cell).strip()
                    if v:
                        empty = False
            vals.append(v)

        if empty or not vals[1].strip():  # must have NOTAM #
            continue

        rows.append(vals)

    return rows


def read_xls_rows(file_bytes: bytes) -> List[List[str]]:
    if not xlrd:
        raise ImportError("xlrd is required to parse .xls (pip install xlrd==1.2.0)")

    book = xlrd.open_workbook(file_contents=file_bytes)
    sheet = book.sheet_by_index(0)

    header_row_idx = None
    col_map = None

    # 1️⃣ Find header row
    for i in range(min(30, sheet.nrows)):
        values = [sheet.cell_value(i, j) for j in range(sheet.ncols)]
        col_map = _find_header_map(values)
        if col_map:
            header_row_idx = i
            break

    if header_row_idx is None:
        raise ValueError("Unable to locate NOTAM header row in XLS file")

    # 2️⃣ Parse rows
    rows = []
    for i in range(header_row_idx + 1, sheet.nrows):
        vals = []
        empty = True
        for exp_idx in range(7):
            col_idx = col_map.get(exp_idx)
            v = ""
            if col_idx is not None and col_idx < sheet.ncols:
                v = str(sheet.cell_value(i, col_idx)).strip()
                if v:
                    empty = False
            vals.append(v)

        if empty or not vals[1].strip():
            continue

        rows.append(vals)

    return rows


def import_xls_bytes_to_notams(file_bytes: bytes, filename: str) -> int:
    name = (filename or "").lower()

    # Detect file type
    is_xlsx = name.endswith(".xlsx") or file_bytes[:2] == b"PK"
    is_xls = name.endswith(".xls") or file_bytes[:8] == b"\xD0\xCF\x11\xE0\xA1\xB1\x1A\xE1"

    if is_xlsx:
        recs = read_xlsx_rows(file_bytes)
    elif is_xls:
        recs = read_xls_rows(file_bytes)
    else:
        # Try XLSX first
        try:
            recs = read_xlsx_rows(file_bytes)
        except Exception:
            recs = read_xls_rows(file_bytes)

    cnt = 0
    now = datetime.utcnow().isoformat() + "Z"

    for vals in recs:
        try:
            insert_notam(
                location=vals[0].strip().upper(),
                notam_lta_number=vals[1].strip(),
                klass=vals[2].strip(),
                issue_date_utc=vals[3].strip(),
                effective_date_utc=vals[4].strip(),
                expiration_date_utc=vals[5].strip(),
                notam_condition_subject_title=vals[6].strip(),
                raw_text=vals[6].strip(),
                uploaded_at=now,
            )
            cnt += 1
        except Exception:
            continue

    return cnt
